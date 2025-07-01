import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
import threading

class ICTReportConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("ICT Report Converter")
        self.root.geometry("600x500")
        self.root.resizable(True, True)
        
        # Variables
        self.txt_files = []
        self.output_directory = tk.StringVar()
        self.program_var = tk.StringVar(value="AUT1077-HC00")
        self.smt_run_var = tk.StringVar(value=datetime.now().strftime('%d-%b-%y'))
        self.smt_line_var = tk.StringVar(value="Line 9")
        
        self.create_widgets()
        self.center_window()
    
    def center_window(self):
        """Center the window on the screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """Create and arrange all GUI widgets"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="ICT Report Converter", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Header Information Section
        header_frame = ttk.LabelFrame(main_frame, text="Product Information", padding="10")
        header_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        header_frame.columnconfigure(1, weight=1)
        
        # Program
        ttk.Label(header_frame, text="Program:").grid(row=0, column=0, sticky=tk.W, pady=5)
        program_entry = ttk.Entry(header_frame, textvariable=self.program_var, width=30)
        program_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # SMT Run
        ttk.Label(header_frame, text="SMT Run:").grid(row=1, column=0, sticky=tk.W, pady=5)
        smt_run_entry = ttk.Entry(header_frame, textvariable=self.smt_run_var, width=30)
        smt_run_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # SMT Line
        ttk.Label(header_frame, text="SMT Line:").grid(row=2, column=0, sticky=tk.W, pady=5)
        smt_line_entry = ttk.Entry(header_frame, textvariable=self.smt_line_var, width=30)
        smt_line_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # File Selection Section
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        file_frame.columnconfigure(0, weight=1)
        
        # TXT Files Selection
        txt_frame = ttk.Frame(file_frame)
        txt_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        txt_frame.columnconfigure(0, weight=1)
        
        ttk.Button(txt_frame, text="Select TXT Files", 
                  command=self.select_txt_files).grid(row=0, column=0, sticky=tk.W)
        
        self.txt_files_label = ttk.Label(txt_frame, text="No files selected", 
                                        foreground="gray")
        self.txt_files_label.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        
        # Output Directory Selection
        output_frame = ttk.Frame(file_frame)
        output_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(15, 5))
        output_frame.columnconfigure(1, weight=1)
        
        ttk.Button(output_frame, text="Select Output Folder", 
                  command=self.select_output_directory).grid(row=0, column=0, sticky=tk.W)
        
        self.output_label = ttk.Label(output_frame, textvariable=self.output_directory, 
                                     foreground="gray")
        self.output_label.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))
        
        # Progress Section
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress_var = tk.StringVar(value="Ready to process files...")
        self.progress_label = ttk.Label(progress_frame, textvariable=self.progress_var)
        self.progress_label.grid(row=0, column=0, sticky=tk.W)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(0, 10))
        
        self.convert_button = ttk.Button(button_frame, text="Convert Files", 
                                        command=self.start_conversion, style='Accent.TButton')
        self.convert_button.grid(row=0, column=0, padx=(0, 10))
        
        ttk.Button(button_frame, text="Exit", 
                  command=self.root.quit).grid(row=0, column=1)
        
        # Log Area
        log_frame = ttk.LabelFrame(main_frame, text="Conversion Log", padding="10")
        log_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
        # Text widget with scrollbar
        text_frame = ttk.Frame(log_frame)
        text_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        self.log_text = tk.Text(text_frame, height=8, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
    
    def select_txt_files(self):
        """Open file dialog to select TXT files"""
        files = filedialog.askopenfilenames(
            title="Select TXT Files",
            filetypes=[("Text files", "*.txt *.TXT"), ("All files", "*.*")]
        )
        
        if files:
            self.txt_files = list(files)
            file_count = len(files)
            if file_count == 1:
                self.txt_files_label.config(text=f"1 file selected: {os.path.basename(files[0])}")
            else:
                self.txt_files_label.config(text=f"{file_count} files selected")
            self.txt_files_label.config(foreground="black")
        else:
            self.txt_files = []
            self.txt_files_label.config(text="No files selected", foreground="gray")
    
    def select_output_directory(self):
        """Open dialog to select output directory"""
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_directory.set(directory)
            self.output_label.config(foreground="black")
        else:
            self.output_directory.set("")
            self.output_label.config(foreground="gray")
    
    def log_message(self, message):
        """Add message to log area"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def parse_csv_report(self, file_path):
        """Parse the CSV report file and return a cleaned DataFrame"""
        try:
            df = pd.read_csv(file_path, skipinitialspace=True)
            df.columns = df.columns.str.strip()
            
            expected_columns = ['Result', 'Board', 'Type', 'Part', 'ActVal', 'StdVal', 'HL', 'LL', 'Mode', 'Range', 'TestVal']
            missing_columns = [col for col in expected_columns if col not in df.columns]
            
            if missing_columns:
                self.log_message(f"Warning: Missing columns in {os.path.basename(file_path)}: {missing_columns}")
            
            return df
            
        except Exception as e:
            self.log_message(f"Error reading CSV file {os.path.basename(file_path)}: {e}")
            return None
    
    def create_formatted_excel(self, df, product_info, output_path):
        """Create a formatted Excel file matching the template exactly"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Report"
        ws.sheet_view.showGridLines = False
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True)
        data_font = Font(name='Arial', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        # Title
        ws['A1'] = 'Product Information'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:B1')
        
        # Product Information Section
        info_data = [
            ['Program', product_info['Program']],
            ['SMT Run', product_info['SMT Run']],
            ['SMT Line', product_info['SMT Line']]
        ]
        
        for i, (key, value) in enumerate(info_data, start=2):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = header_font
            ws[f'A{i}'].fill = header_fill
            ws[f'A{i}'].border = thin_border
            ws[f'A{i}'].alignment = center_alignment
            ws[f'B{i}'].font = data_font
            ws[f'B{i}'].border = thin_border
            ws[f'B{i}'].alignment = center_alignment
        
        # Legend Section
        legend_data = [
            [('Result', 'Result of the measured', 'HL', 'High Rate Tolerance')],
            [('Board', 'Module of the board', 'LL', 'Low Rate Tolerance')],
            [('Type', 'Type of component', 'Mode', 'Type of measure: CC - Constant current,\nAC - Alternate current,\nLV - Low voltage')],
            [('Part', 'Name of the component', 'Range', 'Range of measured')],
            [('ActVal', 'Value paralel\ncomponents', 'TestVal', 'Real measured value')],
            [('StdVal', 'BOM value components', '', '')]
        ]
        
        for row_idx, row_data in enumerate(legend_data, start=7):
            for col_idx, cell_value in enumerate(row_data[0]):
                col_letter = chr(65 + col_idx)
                cell = ws[f'{col_letter}{row_idx}']
                cell.value = cell_value
                cell.border = thin_border
                cell.alignment = center_alignment
                
                if col_idx in [0, 2]:  # Headers (A and C columns)
                    cell.font = header_font
                    cell.fill = header_fill
                else:  # Descriptions (B and D columns)
                    cell.font = data_font
        
        # Merge last row for StdVal
        ws.merge_cells('C12:D12')
        
        # Data section
        data_start_row = 14
        headers = ['Result', 'Board', 'Type', 'Part', 'ActVal', 'StdVal', 'HL', 'LL', 'Mode', 'Range', 'TestVal']
        
        for j, header in enumerate(headers):
            col_letter = chr(65 + j)
            cell = ws[f'{col_letter}{data_start_row}']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Add data rows
        for i, row in df.iterrows():
            row_num = data_start_row + 1 + i
            result = str(row['Result']).strip().upper()
            
            for j, header in enumerate(headers):
                col_letter = chr(65 + j)
                cell = ws[f'{col_letter}{row_num}']
                cell.value = row[header] if header in row else ''
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_alignment
                
                if j == 0 and result == 'PASS':
                    cell.fill = pass_fill
        
        # Adjust column widths
        column_widths = {'A': 8, 'B': 8, 'C': 12, 'D': 15, 'E': 10, 'F': 10, 'G': 8, 'H': 8, 'I': 8, 'J': 8, 'K': 12}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(output_path)
    
    def start_conversion(self):
        """Start the conversion process in a separate thread"""
        if not self.txt_files:
            messagebox.showerror("Error", "Please select TXT files to convert.")
            return
        
        if not self.output_directory.get():
            messagebox.showerror("Error", "Please select an output directory.")
            return
        
        if not all([self.program_var.get().strip(), self.smt_run_var.get().strip(), self.smt_line_var.get().strip()]):
            messagebox.showerror("Error", "Please fill in all product information fields.")
            return
        
        # Disable convert button during processing
        self.convert_button.config(state='disabled')
        
        # Clear log
        self.log_text.delete(1.0, tk.END)
        
        # Start conversion in separate thread
        thread = threading.Thread(target=self.convert_files)
        thread.daemon = True
        thread.start()
    
    def convert_files(self):
        """Convert all selected files"""
        try:
            total_files = len(self.txt_files)
            processed_count = 0
            failed_count = 0
            
            self.progress_bar.config(maximum=total_files)
            
            product_info = {
                'Program': self.program_var.get().strip(),
                'SMT Run': self.smt_run_var.get().strip(),
                'SMT Line': self.smt_line_var.get().strip()
            }
            
            for i, txt_file in enumerate(self.txt_files):
                try:
                    filename = os.path.basename(txt_file)
                    self.progress_var.set(f"Processing: {filename}")
                    self.log_message(f"Processing: {filename}")
                    
                    # Parse the CSV data
                    df = self.parse_csv_report(txt_file)
                    
                    if df is None:
                        self.log_message(f"Failed to read data from {filename}")
                        failed_count += 1
                        continue
                    
                    self.log_message(f"Successfully read {len(df)} records.")
                    
                    # Create output filename
                    base_name = os.path.splitext(filename)[0]
                    output_file = os.path.join(self.output_directory.get(), f"Test_Report_{base_name}.xlsx")
                    
                    # Create the formatted Excel file
                    self.create_formatted_excel(df, product_info, output_file)
                    
                    # Display summary statistics
                    result_counts = df['Result'].value_counts()
                    self.log_message(f"Summary for {filename}:")
                    for result, count in result_counts.items():
                        self.log_message(f"  {result}: {count}")
                    
                    self.log_message(f"✓ Saved: {os.path.basename(output_file)}")
                    processed_count += 1
                    
                except Exception as e:
                    self.log_message(f"✗ Error processing {filename}: {e}")
                    failed_count += 1
                
                # Update progress
                self.progress_bar.config(value=i + 1)
                self.root.update_idletasks()
            
            # Final summary
            self.progress_var.set("Conversion completed!")
            self.log_message("\n" + "="*50)
            self.log_message("BATCH PROCESSING COMPLETE")
            self.log_message("="*50)
            self.log_message(f"Successfully processed: {processed_count} files")
            self.log_message(f"Failed to process: {failed_count} files")
            self.log_message(f"Output directory: {self.output_directory.get()}")
            
            # Show completion message
            if processed_count > 0:
                messagebox.showinfo("Success", f"Conversion completed!\n\nProcessed: {processed_count} files\nFailed: {failed_count} files")
            else:
                messagebox.showerror("Error", "No files were successfully processed.")
                
        except Exception as e:
            self.log_message(f"Unexpected error during conversion: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
        
        finally:
            # Re-enable convert button
            self.convert_button.config(state='normal')

def main():
    root = tk.Tk()
    app = ICTReportConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
